"""
Author: Rayla Kurosaki

GitHub: https://github.com/rkp1503
"""

import json
import shutil

import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook


def create_backup_files() -> None:
    src: str = f"assets/data/database.json"
    dst: str = f"assets/data/database - Backup.json"
    shutil.copyfile(src, dst)

    src: str = f"assets/data/database.xlsx"
    dst: str = f"assets/data/database - Backup.xlsx"
    shutil.copyfile(src, dst)
    return None


def populate_local_database() -> dict[str, dict]:
    with open("assets/data/database.json", "r") as file:
        json_data: dict = json.load(file)
        pass
    return json_data


def update_database_from_excel(database: dict) -> None:
    sheet_names: list[str] = ["Characters", "Weapons"]
    for sheet_name in sheet_names:
        worksheet = load_workbook("assets/data/database.xlsx")[sheet_name]
        for (i, row) in enumerate(worksheet.iter_rows()):
            if i > 0:
                name: str = row[0].value
                excel_data: list = [
                    row[i].value for i in range(worksheet.max_column)
                ]
                if name in database[sheet_name].keys():
                    json_data: list = list(database[sheet_name][name].values())
                    if not excel_data == json_data:
                        json_keys: list = database[sheet_name][name].keys()
                        database[sheet_name][name] = dict(zip(json_keys,
                                                              excel_data))
                        pass
                    pass
                pass
            pass
        pass
    return None


def web_scraper_character(url: str) -> dict:
    with open("assets/data/misc.json", "r") as file:
        misc_data: dict = json.load(file)
        pass
    character: dict = {
        "Name": None,
        "Quality": None,
        "Sex": None,
        "Element": None,
        "Nation": None,
        "Weapon": None,
        "Model Type": None,
        "Signature Weapon": None,
        "Base HP": None,
        "Base ATK": None,
        "Base DEF": None,
        "Ascension Stat": None,
        "Release Date": None,
        "Sort Order": None
    }
    soup_1 = BeautifulSoup(requests.get(url).content, "html.parser")
    result = soup_1.find("aside")
    for h2 in result.find_all("h2"):
        # Get Character Name
        if h2.has_attr("data-source"):
            name: str = h2.get_text().strip()
            character["Name"] = name
            pass
        pass
    for td in result.find_all("td"):
        if td.has_attr("data-source"):
            soup_2 = BeautifulSoup(td.prettify(), "html.parser")
            # Get Quality
            if td.get("data-source") == "quality":
                quality: str = soup_2.find("img").get("title").split(" ")[0]
                character["Quality"] = int(quality)
                pass
            # Get Weapon
            elif td.get("data-source") == "weapon":
                weapon_type: str = soup_2.find("a").get("title").strip()
                character["Weapon"] = weapon_type
                pass
            # Get Element
            elif td.get("data-source") == "element":
                if character["Name"] != "Traveler":
                    element: str = soup_2.find("a").get("title").strip()
                    character["Element"] = element
                    pass
                pass
            pass
        # Get Sex and Model Type
        else:
            if character["Name"] != "Traveler":
                model_type: str = td.get_text()
                character["Model Type"] = model_type
                sex: str = model_type.split(" ")[-1]
                character["Sex"] = sex
                pass
            pass
        pass
    field: str = "pi-item pi-panel pi-border-color wds-tabber"
    section = result.find("section", class_=field)
    res_4 = section.find("div", class_="wds-tab__content wds-is-current")
    for res_5 in res_4.find_all("div"):
        # Get Nation
        if res_5.get("data-source") == "region":
            soup_3 = BeautifulSoup(res_5.prettify(), "html.parser")
            nation: str = soup_3.find("div").find("div").get_text().strip()
            nation = nation.replace("\n", "|").split("|")[0].strip()
            character["Nation"] = nation
            pass
        # Get Release Date
        elif res_5.get("data-source") == "releaseDate":
            soup_3 = BeautifulSoup(res_5.prettify(), "html.parser")
            res_6 = soup_3.find("div").find("div").get_text(separator="|")
            month, day, year = (res_6.strip().split("|")[0]).strip().split(" ")
            month = misc_data["Months"][month]
            day = day.split(",")[0]
            character["Release Date"] = f"{year}/{month}/{day}"
            pass
        pass
    table = soup_1.find("table", class_="wikitable ascension-stats")
    results = table.find_all("tr")
    for (i, res_7) in enumerate(results):
        soup_4 = BeautifulSoup(res_7.prettify(), "html.parser")
        # Get Ascension Stat
        if i == 0:
            res_8 = soup_4.find_all("th")
            soup_5 = BeautifulSoup(res_8[-1].prettify(), "html.parser")
            result = soup_5.find("span").find("a")
            ascension_stat: str = result.get_text().strip()
            character["Ascension Stat"] = ascension_stat
            pass
        # Get Base Stats
        elif i == 1:
            td_s = soup_4.find_all("td")
            # Base HP
            base_hp: str = td_s[2].get_text().strip().replace(',', '')
            character["Base HP"] = float(base_hp)
            # Base ATK
            base_atk: str = td_s[3].get_text().strip().replace(',', '')
            character["Base ATK"] = float(base_atk)
            # Base DEF
            base_def: str = td_s[4].get_text().strip().replace(',', '')
            character["Base DEF"] = float(base_def)
            pass
        none_lst: list[str | None] = [
            character["Ascension Stat"], character["Base HP"],
            character["Base ATK"], character["Base DEF"]
        ]
        if None not in none_lst:
            break
            pass
        pass
    return character


def character_data(database: dict) -> None:
    with open("assets/data/urls.json", "r") as file:
        urls_data: dict = json.load(file)
        pass
    soup_1 = BeautifulSoup(requests.get(urls_data["Characters"]).content,
                           "html.parser")
    result = soup_1.find("div", class_="mw-body-content mw-content-ltr")
    result = result.find("div", class_="mw-parser-output")
    result = result.find("table")
    for (i, row) in enumerate(result.find_all("tr")):
        if i > 0:
            soup_2 = BeautifulSoup(row.prettify(), "html.parser")
            td = soup_2.find_all("td")[1]
            soup_3 = BeautifulSoup(td.prettify(), "html.parser")
            res_1 = soup_3.find("a")
            name: str = res_1.get_text().strip()
            if name not in database["Characters"].keys():
                print(f"\t\tGetting {name}'s data...")
                character_url: str = res_1.get("href")
                url: str = f"{urls_data['Main']}{character_url}"
                character: dict = web_scraper_character(url)
                database["Characters"][name] = character
                pass
            pass
        pass
    return None


def web_scraper_weapon(url: str) -> dict:
    with open("assets/data/misc.json", "r") as file:
        misc_data: dict = json.load(file)
        pass
    weapon: dict = {
        "Name": None,
        "Quality": None,
        "Weapon Type": None,
        "Base ATK": None,
        "Sub Stat": None,
        "Sub Stat Val": None,
        "Release Date": None,
        "Sort Order": None
    }
    soup_1 = BeautifulSoup(requests.get(url).content, "html.parser")
    result = soup_1.find("aside")
    # Get Name
    weapon_name: str = result.find("h2").get_text().replace("\\xad", "")
    if weapon_name not in misc_data["Ignore_Weapons"]:
        weapon["Name"] = weapon_name
        field: str = "pi-item pi-data pi-item-spacing pi-border-color"
        for div_i_1 in result.find_all("div", class_=field):
            data_type: str = div_i_1.get("data-source").strip()
            # Get Type
            if data_type == "type":
                soup_2 = BeautifulSoup(div_i_1.prettify(), "html.parser")
                weapon_type: str = soup_2.find_all("a")[-1].get_text().strip()
                weapon["Weapon Type"] = weapon_type
                pass
            # Get Quality
            elif data_type == "quality":
                quality: str = div_i_1.find("img").get("title").split(" ")[0]
                weapon["Quality"] = int(quality)
                pass
            # Get Release Date
            elif data_type == "releaseDate":
                release_date: str = div_i_1.find("div").get_text(separator="|")
                month, day, year = release_date.split("|")[0].split(" ")
                month = misc_data["Months"][month]
                day = day.split(",")[0]
                weapon["Release Date"] = f"{year}/{month}/{day}"
                pass
            pass
        field: str = "pi-item pi-smart-group pi-border-color"
        result = soup_1.find("section", class_=field)
        field: str = "pi-smart-group-body pi-border-color"
        result = result.find_all("section", class_=field)[-1]
        for (i, div_i_2) in enumerate(result.find_all("div")):
            data: str = div_i_2.get_text()
            # Get Max Base ATK
            if i == 0:
                if data != "Unknown":
                    weapon["Base ATK"] = int(data.split(" - ")[-1])
                    pass
                pass
            # Get Sub Stat
            elif i == 1:
                weapon["Sub Stat"] = data
                pass
            # Get Sub Stat Value
            elif i == 2:
                sub_stat_val: str = data.split(" - ")[-1]
                if "%" in sub_stat_val:
                    weapon["Sub Stat Val"] = float(sub_stat_val[:-1]) / 100
                    pass
                else:
                    weapon["Sub Stat Val"] = float(sub_stat_val)
                    pass
                pass
            pass
        pass
    return weapon


def weapon_data(database: dict) -> None:
    with open("assets/data/urls.json", "r") as file:
        urls_data: dict = json.load(file)
        pass
    with open("assets/data/misc.json", "r") as file:
        misc_data: dict = json.load(file)
        pass
    for (weapon_type, url) in urls_data["Weapons"].items():
        soup = BeautifulSoup(requests.get(url).content, "html.parser")
        field: str = "nowraplinks mw-collapsible mw-collapsed navbox-inner"
        result = soup.find("table", class_=field)
        for (i, a_i) in enumerate(result.find_all("a")):
            if i > 0:
                weapon_url_end: str = a_i.get("href")
                url: str = f"{urls_data['Main']}{weapon_url_end}"
                if url not in urls_data["Weapons"].values():
                    name: str = a_i.get("title")
                    if name not in database["Weapons"].keys():
                        if name not in misc_data["Ignore_Weapons"]:
                            print(f"\t\tGetting {name}'s data...")
                            weapon: dict = web_scraper_weapon(url)
                            database["Weapons"][name] = weapon
                            pass
                        pass
                    pass
                pass
            pass
        pass
    return None


def update_excel_file(database: dict) -> None:
    workbook = load_workbook("assets/data/database.xlsx")
    sheet_names: list[str] = ["Characters", "Weapons"]
    for sheet_name in sheet_names:
        for (name, data) in database[sheet_name].items():
            if name not in database[sheet_name].keys():
                workbook[sheet_name].append(list(data.values()))
                pass
            pass
        pass
    workbook.save("assets/data/database.xlsx")
    return None


def update_json_file(database: dict) -> None:
    with open("assets/data/database.json", "w") as file:
        json.dump(database, file, indent=2)
        pass
    return None


def main() -> None:
    print("Creating backup files...")
    create_backup_files()

    print("Populating local database...")
    database: dict = populate_local_database()

    print("Updating database from Excel file...")
    update_database_from_excel(database)

    print("Updating database from global source...")
    character_data(database)
    weapon_data(database)

    print("Updating Excel file...")
    update_excel_file(database)

    print("Updating json file...")
    update_json_file(database)
    return None


if __name__ == '__main__':
    main()
    pass
