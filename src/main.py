"""
Author: Rayla Kurosaki

GitHub: https://github.com/rkp1503
"""

import json
import shutil

import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook

from Character import Character
from Weapon import Weapon

MAIN_URL: str = "https://genshin-impact.fandom.com"

WEAPONS_URLS: dict = {
    "Swords": "https://genshin-impact.fandom.com/wiki/Sword",
    "Claymore": "https://genshin-impact.fandom.com/wiki/Claymore",
    "Polearm": "https://genshin-impact.fandom.com/wiki/Polearm",
    "Catalyst": "https://genshin-impact.fandom.com/wiki/Catalysts",
    "Bow": "https://genshin-impact.fandom.com/wiki/Bow"
}

IGNORE_WEAPONS: list = [
    "Prized Isshin Blade (Awakened)", "Prized Isshin Blade (Shattered)",
    "Sword of Narzissenkreuz (Quest)", "Quartz", "Amber Catalyst", "Ebony Bow",
]

CHARACTERS_URL: str = "https://genshin-impact.fandom.com/wiki/Character/List"

MONTH_CONVERSION_FULL: dict = {
    "January": "01", "February": "02", "March": "03", "April": "04",
    "May": "05", "June": "06", "July": "07", "August": "08", "September": "09",
    "October": "10", "November": "11", "December": "12"
}

COLORS: dict = {
    "Quality":
        {
            "3": "#C24F54",
            "4": "#8B73B4",
            "5": "#C78853"
        }
    ,
    "Vision":
        {
            "Anemo": "#33CCB3",
            "Geo": "#CFA726",
            "Electro": "#CC72E8",
            "Dendro": "#7BB42D",
            "Hydro": "#1C72FD",
            "Pyro": "#E0311D",
            "Cryo": "#98C8E8"
        }
    ,
    "Nation":
        {
            "Mondstadt": "#33CCB3",
            "Liyue": "#CFA726",
            "Inazuma": "#CC72E8",
            "Sumeru": "#7BB42D",
            "Fontaine": "#1C72FD",
            "Natlan": "#E0311D",
            "Snezhnaya": "#98C8E8"
        }
    ,
    "DMG Bonus":
        {
            "Anemo DMG Bonus": "#33CCB3",
            "Geo DMG Bonus": "#CFA726",
            "Electro DMG Bonus": "#CC72E8",
            "Dendro DMG Bonus": "#7BB42D",
            "Hydro DMG Bonus": "#1C72FD",
            "Pyro DMG Bonus": "#E0311D",
            "Cryo DMG Bonus": "#98C8E8"
        }
}


def populate_json_file() -> dict:
    database: dict = {"Characters": {}, "Weapons": {}}
    with open("database.json", "r") as file:
        json_data: dict = json.load(file)
        pass
    for (data_type, data_db) in json_data.items():
        for data in data_db.values():
            if data_type == "Characters":
                character: Character = Character()
                character.set_name(data["name"])
                character.set_quality(data["quality"])
                character.set_vision(data["vision"])
                character.set_sex(data["sex"])
                character.set_nation(data["nation"])
                character.set_weapon(data["weapon"])
                character.set_signature_weapon(data["signature_weapon"])
                character.set_ascension_stat(data["ascension_stat"])
                character.set_ascension_stat_val(data["ascension_stat_val"],
                                                 local=True)
                character.set_release_date(data["release_date"])
                character.set_sort_order(data["sort_order"])
                database[data_type][character.get_name()] = character
                pass
            else:
                weapon: Weapon = Weapon()
                weapon.set_name(data["name"])
                weapon.set_quality(data["quality"])
                weapon.set_weapon_type(data["weapon_type"])
                weapon.set_base_atk(data["base_atk"])
                weapon.set_sub_stat(data["sub_stat"])
                weapon.set_sub_stat_val(data["sub_stat_val"], local=True)
                weapon.set_release_date(data["release_date"])
                weapon.set_sort_order(data["sort_order"])
                weapon.set_refinement(data["refinement"])
                weapon.set_level(data["level"])
                database[data_type][weapon.get_name()] = weapon
                pass
            pass
        pass
    return database


def update_database_from_excel(database: dict) -> None:
    worksheet = load_workbook("database.xlsx")["Characters"]
    for (i, row) in enumerate(worksheet.iter_rows()):
        if i > 0:
            name: str = row[0].value
            signature_weapon: str | None = row[6].value
            sort_order: int | None = int(row[10].value)
            if name in database["Characters"].keys():
                character: Character = database["Characters"][name]
                if signature_weapon is not None:
                    character.set_signature_weapon(signature_weapon)
                    pass
                if sort_order is not None:
                    character.set_sort_order(sort_order)
                    pass
                pass
            pass
        pass
    worksheet = load_workbook("database.xlsx")["Weapons"]
    for (i, row) in enumerate(worksheet.iter_rows()):
        if i > 0:
            name: str = row[0].value
            sort_order: int | None = int(row[7].value)
            refinement: int | None = int(row[8].value)
            level: int | None = int(row[9].value)
            if name in database["Weapons"].keys():
                weapon: Weapon = database["Weapons"][name]
                if sort_order is not None:
                    weapon.set_sort_order(sort_order)
                    pass
                if refinement is not None:
                    weapon.set_refinement(refinement)
                    pass
                if level is not None:
                    weapon.set_level(level)
                    pass
                pass
            pass
        pass
    return None


def web_scraper_character(url: str) -> Character:
    character: Character = Character()
    soup_1 = BeautifulSoup(requests.get(url).content, "html.parser")
    result = soup_1.find("aside")
    for h2 in result.find_all("h2"):
        # Get Name
        if h2.has_attr("data-source"):
            name: str = h2.get_text().strip()
            character.set_name(name)
            pass
        pass
    for td in result.find_all("td"):
        if td.has_attr("data-source"):
            soup_2 = BeautifulSoup(td.prettify(), "html.parser")
            # Get Quality
            if td.get("data-source") == "quality":
                stars: int = int(soup_2.find("img").get("title").split(" ")[0])
                character.set_quality(stars)
                pass
            # Get Weapon
            elif td.get("data-source") == "weapon":
                weapon_type: str = soup_2.find("a").get("title").strip()
                character.set_weapon(weapon_type)
                pass
            # Get Element
            elif td.get("data-source") == "element":
                if character.get_name() != "Traveler":
                    vision: str = soup_2.find("a").get("title").strip()
                    character.set_vision(vision)
                    pass
                pass
            pass
        # Get Model Type and Sex
        else:
            if character.get_name() != "Traveler":
                sex: str = td.get_text().split(" ")[-1]
                character.set_sex(sex)
                pass
            pass
        pass
    field: str = "pi-item pi-panel pi-border-color wds-tabber"
    section = result.find("section", class_=field)
    res_4 = section.find("div", class_="wds-tab__content wds-is-current")
    for res_5 in res_4.find_all("div"):
        # Get region
        if res_5.get("data-source") == "region":
            soup_3 = BeautifulSoup(res_5.prettify(), "html.parser")
            nation: str = soup_3.find("div").find("div").get_text().strip()
            nation = nation.replace("\n", "|").split("|")[0].strip()
            character.set_nation(nation)
            pass
        # Get Release Date
        elif res_5.get("data-source") == "releaseDate":
            soup_3 = BeautifulSoup(res_5.prettify(), "html.parser")
            res_6 = soup_3.find("div").find("div").get_text(separator="|")
            month, day, year = (res_6.strip().split("|")[0]).strip().split(" ")
            month = MONTH_CONVERSION_FULL[month]
            day = day.split(",")[0]
            character.set_release_date(f"{year}/{month}/{day}")
            pass
        pass
    table = soup_1.find("table", class_="wikitable ascension-stats")
    results = table.find_all("tr")
    for (i, res_7) in enumerate(results):
        soup_4 = BeautifulSoup(res_7.prettify(), "html.parser")
        # Get Ascension Stat
        if i == 0:
            res_8 = soup_4.find_all("th")
            for (j, col) in enumerate(res_8):
                if j == len(res_8) - 1:
                    soup_5 = BeautifulSoup(col.prettify(), "html.parser")
                    result = soup_5.find("span").find("a")
                    ascension_stat: str = result.get_text().strip()
                    character.set_ascension_stat(ascension_stat)
                    pass
                pass
            pass
        # Get Ascension Stat Val
        elif i == len(results) - 2:
            td_s = soup_4.find_all("td")
            for (j, res_11) in enumerate(td_s):
                if j == len(td_s) - 1:
                    ascension_stat_value: str = res_11.get_text().strip()
                    character.set_ascension_stat_val(ascension_stat_value)
                    pass
                pass
            pass
        pass
    return character


def character_data(database: dict) -> None:
    soup_1 = BeautifulSoup(requests.get(CHARACTERS_URL).content, "html.parser")
    result = soup_1.find("div", class_="mw-body-content mw-content-ltr")
    result = result.find("div", class_="mw-parser-output")
    result = result.find("table")
    for (i, row) in enumerate(result.find_all("tr")):
        if i > 0:
            soup_2 = BeautifulSoup(row.prettify(), "html.parser")
            td = soup_2.find_all("td")[1]
            soup_3 = BeautifulSoup(td.prettify(), "html.parser")
            res_1 = soup_3.find("a")
            name: str = res_1.get_text().strip()
            if name not in database["Characters"].keys():
                print(f"\t\tGetting {name}'s data...")
                character_url: str = res_1.get("href")
                url: str = f"{MAIN_URL}{character_url}"
                character: Character = web_scraper_character(url)
                database["Characters"][name] = character
                pass
            pass
        pass
    return None


def web_scraper_weapon(url: str) -> Weapon:
    weapon: Weapon = Weapon()
    soup_1 = BeautifulSoup(requests.get(url).content, "html.parser")
    result = soup_1.find("aside")
    # Get Name
    weapon_name: str = result.find("h2").get_text()
    if weapon_name not in IGNORE_WEAPONS:
        weapon.set_name(weapon_name)
        field: str = "pi-item pi-data pi-item-spacing pi-border-color"
        for div_i_1 in result.find_all("div", class_=field):
            data_type: str = div_i_1.get("data-source").strip()
            # Get Type
            if data_type == "type":
                soup_2 = BeautifulSoup(div_i_1.prettify(), "html.parser")
                weapon_type: str = soup_2.find_all("a")[-1].get_text().strip()
                weapon.set_weapon_type(weapon_type)
                pass
            # Get Quality
            elif data_type == "quality":
                stars: int = int(
                    div_i_1.find("img").get("title").split(" ")[0])
                weapon.set_quality(stars)
                pass
            # Get Release Date
            elif data_type == "releaseDate":
                release_date: str = div_i_1.find("div").get_text(separator="|")
                month, day, year = release_date.split("|")[0].split(" ")
                month = MONTH_CONVERSION_FULL[month]
                day = day.split(",")[0]
                weapon.set_release_date(f"{year}/{month}/{day}")
                pass
            pass
        field: str = "pi-item pi-smart-group pi-border-color"
        result = soup_1.find("section", class_=field)
        field: str = "pi-smart-group-body pi-border-color"
        result = result.find_all("section", class_=field)[-1]
        for (i, div_i_2) in enumerate(result.find_all("div")):
            data: str = div_i_2.get_text()
            # Get Max Base ATK
            if i == 0:
                weapon.set_base_atk(int(data.split(" - ")[-1]))
                pass
            # Get Sub Stat
            elif i == 1:
                weapon.set_sub_stat(data)
                pass
            # Get Sub Stat Value
            elif i == 2:
                weapon.set_sub_stat_val(data.split(" - ")[-1])
                pass
            pass
        pass
    return weapon


def weapon_data(database: dict) -> None:
    for (weapon_type, url) in WEAPONS_URLS.items():
        soup = BeautifulSoup(requests.get(url).content, "html.parser")
        field: str = "nowraplinks mw-collapsible mw-collapsed navbox-inner"
        result = soup.find("table", class_=field)
        for (i, a_i) in enumerate(result.find_all("a")):
            if i > 0:
                weapon_url_end: str = a_i.get("href")
                url: str = f"{MAIN_URL}{weapon_url_end}"
                if url not in WEAPONS_URLS.values():
                    name: str = a_i.get("title")
                    if name not in database["Weapons"].keys():
                        if name not in IGNORE_WEAPONS:
                            print(f"\t\tGetting {name}'s data...")
                            weapon: Weapon = web_scraper_weapon(url)
                            weapon.set_name(name)
                            database["Weapons"][name] = weapon
                        pass
                    pass
                pass
            pass
        pass
    return None


def update_excel_file(database: dict) -> None:
    workbook = load_workbook("database.xlsx")
    worksheet = workbook["Characters"]
    names: list = []
    for row in worksheet:
        names.append(row[0].value)
        pass
    for character in database["Characters"].values():
        new_character: str = character.get_name()
        if new_character not in names:
            data: list = [
                character.get_name(), character.get_quality(),
                character.get_vision(), character.get_sex(),
                character.get_nation(), character.get_weapon(),
                character.get_signature_weapon(),
                character.get_ascension_stat(),
                character.get_ascension_stat_val(),
                character.get_release_date(), character.get_sort_order()
            ]
            worksheet.append(data)
            pass
        pass
    worksheet = workbook["Weapons"]
    names: list = []
    for row in worksheet:
        names.append(row[0].value)
        pass
    for weapon in database["Weapons"].values():
        new_weapon: str = weapon.get_name()
        if new_weapon not in names:
            data: list = [
                weapon.get_name(), weapon.get_quality(),
                weapon.get_weapon_type(), weapon.get_base_atk(),
                weapon.get_sub_stat(), weapon.get_sub_stat_val(),
                weapon.get_release_date(), weapon.get_sort_order(),
                weapon.get_refinement(), weapon.get_level()
            ]
            worksheet.append(data)
            pass
        pass
    workbook.save("database.xlsx")
    return None


def update_json_file(database: dict) -> None:
    json_data: dict = {"Characters": {}, "Weapons": {}}
    for (data_type, data_dict) in database.items():
        for (name, data_object) in data_dict.items():
            json_data[data_type][name] = data_object.__dict__
            pass
        pass
    with open("database.json", "w") as file:
        json.dump(json_data, file, indent=2)
        pass
    return None


def create_backup_files() -> None:
    filename: str = "database"
    backup_filename: str = f"{filename} - Backup"
    shutil.copyfile(f"{filename}.xlsx", f"{backup_filename}.xlsx")
    shutil.copyfile(f"{filename}.json", f"{backup_filename}.json")
    return None


def main() -> None:
    print("Creating backup database files...")
    create_backup_files()

    print("Populating database...")
    database: dict = populate_json_file()

    print("Updating database from Excel file...")
    update_database_from_excel(database)

    print("Updating database from global source...")
    character_data(database)
    weapon_data(database)

    print("Updating Excel file...")
    update_excel_file(database)

    print("Updating json file...")
    update_json_file(database)
    return None


if __name__ == '__main__':
    main()
    pass
